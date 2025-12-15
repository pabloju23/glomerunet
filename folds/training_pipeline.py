import os
import sys

# Añadir el directorio padre al sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
# Set environment variable to select only GPU 1 and GPU 2
os.environ["CUDA_VISIBLE_DEVICES"] = "0,1,2"
import tensorflow as tf
# Optimize GPU memory allocation
gpus = tf.config.experimental.list_physical_devices('GPU')
for gpu in gpus:
    tf.config.experimental.set_memory_growth(gpu, True)
import numpy as np
from methods.data_augmentor import create_dataset, create_dataset_with_class_augmentation
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook, load_workbook


# Configuration
IMAGE_SIZE = (512, 512, 3)  # Reduced size to manage memory
BATCH_SIZE = 36  # Further reduced batch size to fit GPU memory
NUM_CLASSES=3

# Train models and save performance summary
performance_summary = []


def update_excel(file_path, row_index, column_values):
    """
    Crea o actualiza un archivo Excel en la fila y columnas especificadas.

    Args:
        file_path (str): Ruta del archivo Excel.
        row_index (int): Índice de la fila (0 para crear un archivo nuevo).
        column_values (dict): Diccionario con los datos a escribir en las columnas.
    """
    # Verificar que column_values sea un diccionario
    if not isinstance(column_values, dict):
        raise ValueError("`column_values` debe ser un diccionario.")
    
    # Crear un archivo nuevo si el índice es 0
    if row_index == 0:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"
        # Escribir encabezados y valores
        headers = list(column_values.keys())
        values = list(column_values.values())
        ws.append(headers)  # Encabezados en la primera fila
        ws.append(values)   # Valores en la segunda fila
        wb.save(file_path)
        print(f"Nuevo archivo Excel creado en: {file_path}")
    else:
        # Verificar si el archivo existe
        if not os.path.exists(file_path):
            print(f"Error: El archivo '{file_path}' no existe.")
            return
        
        # Cargar el archivo existente
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Escribir los valores en la fila indicada
        values = list(column_values.values())
        for col_index, value in enumerate(values, start=1):
            ws.cell(row=row_index + 2, column=col_index, value=value)  # `+1` porque Excel usa base 1
        wb.save(file_path)
        print(f"Valores añadidos en la fila {row_index + 1} del archivo: {file_path}")

def train_and_evaluate(train_img, train_mask, val_img, val_mask, test_img, test_mask, test_fold, val_fold, train_folds, excel_row, callbacks, strategy, base_model=None):
    # Create datasets
    train_dataset = create_dataset_with_class_augmentation(train_img, train_mask, batch_size=BATCH_SIZE, shuffle=True, augmentation=True)
    val_dataset = create_dataset(val_img, val_mask, batch_size=BATCH_SIZE, shuffle=False)
    test_dataset = create_dataset(test_img, test_mask, batch_size=6, shuffle=False)

    with strategy.scope():
        from methods.tversky_metric import OneHotTversky
        target_class_ids=[0, 1, 2]

        # Create and compile model
        iou_metric = tf.keras.metrics.OneHotIoU(name='IoU', num_classes=NUM_CLASSES, target_class_ids=target_class_ids)  

        from methods.dice_metric import OneHotDice
        dice_metric = OneHotDice(num_classes=NUM_CLASSES, name='dsc', target_class_ids=target_class_ids)
        dice_nobg = OneHotDice(num_classes=NUM_CLASSES, name='dsc_nobg', target_class_ids=[1, 2])
        dice_1 = OneHotDice(num_classes=NUM_CLASSES, name='dsc1', target_class_ids= [1])
        dice_2 = OneHotDice(num_classes=NUM_CLASSES, name='dsc2', target_class_ids= [2])
        from methods.tversky_metric import OneHotTversky
        tversky_metric = OneHotTversky(alpha=0.7, beta=0.3, name='tversky', num_classes=NUM_CLASSES, target_class_ids=target_class_ids)

        def categorical_focal_tversky_loss(alpha=0.5, gamma=2.0, tversky_alpha=0.7, tversky_beta=0.3):
            """
            Combina la Focal Loss con la Tversky Loss para problemas de segmentación multiclase.
            Parámetros:
                alpha: Peso para balancear Focal Loss y Tversky Loss.
                gamma: Parámetro de enfoque para la Focal Loss.
                tversky_alpha: Penalización para falsos positivos en Tversky Loss.
                tversky_beta: Penalización para falsos negativos en Tversky Loss.
            """
            def loss(y_true, y_pred):
                # Clipping para evitar errores numéricos
                y_pred = tf.clip_by_value(y_pred, tf.keras.backend.epsilon(), 1.0 - tf.keras.backend.epsilon())
                
                # Tversky Loss
                tp = tf.reduce_sum(y_true * y_pred, axis=[1, 2])
                fp = tf.reduce_sum((1 - y_true) * y_pred, axis=[1, 2])
                fn = tf.reduce_sum(y_true * (1 - y_pred), axis=[1, 2])
                tversky_loss = tf.reduce_mean(1 - (tp / (tp + tversky_alpha * fp + tversky_beta * fn)))
                
                # Focal Categorical Crossentropy
                focal_loss = -y_true * tf.pow(1 - y_pred, gamma) * tf.math.log(y_pred)
                focal_loss = tf.reduce_mean(tf.reduce_sum(focal_loss, axis=-1))
                
                # Combinación de ambas pérdidas
                return alpha * focal_loss + (1 - alpha) * tversky_loss
            
            return loss

        cftl = categorical_focal_tversky_loss(alpha=0.35, gamma=2.0, tversky_alpha=0.7, tversky_beta=0.3)
        cfce = tf.keras.losses.CategoricalFocalCrossentropy(
        # alpha = tf.constant([0.0031,0.1695,0.8274], dtype=tf.float32),    # Peso para las clases minoritarias
        gamma=2,     # Focusing parameter
        ) #   #  tf.keras.losses.BinaryCrossentropy(from_logits=False) #TverskyLoss() #  
  
        initial_lr = 1e-4
        adam_opt = tf.keras.optimizers.Adam(learning_rate=initial_lr)
        opt = adam_opt

        # Create model
        if base_model is not None:
            # Clone the pretrained model to get fresh weights
            model = tf.keras.models.clone_model(base_model)
            model.set_weights(base_model.get_weights())
            print("Model cloned from pretrained base.")
        else:
            from models.keras.keras_deeplab_model import DeeplabV3Plus_largev2
            model = DeeplabV3Plus_largev2(image_size=IMAGE_SIZE[0], num_classes=NUM_CLASSES, 
                                          backbone='xception', weights='imagenet')
            print("New model created from scratch.")
        model.compile(optimizer=opt,
                loss=cftl,
                metrics=[iou_metric, dice_metric, dice_nobg, dice_1, dice_2])
        # Train model
        history = model.fit(train_dataset, epochs=20, 
                            validation_data=val_dataset, 
                            callbacks=callbacks)
        
        model.save(f'paper_checkpoints/deeplab_glomeruli_multiclass_finetunned.keras')

        # Save best epoch loss and metrics
        best_epoch = history.history['val_IoU'].index(max(history.history['val_IoU'])) 
        best_metrics = {
            'epoch': best_epoch + 1,
            'train_loss': history.history['loss'][best_epoch],
            'train_iou': history.history['IoU'][best_epoch],
            'train_dice': history.history['dsc'][best_epoch],
            'train_dice_nobg': history.history['dsc_nobg'][best_epoch],
            'train_dice_2': history.history['dsc2'][best_epoch],
            'val_loss': history.history['val_loss'][best_epoch],
            'val_iou': history.history['val_IoU'][best_epoch],
            'val_dice': history.history['val_dsc'][best_epoch],
            'val_dice_nobg': history.history['val_dsc_nobg'][best_epoch],
            'val_dice_2': history.history['val_dsc2'][best_epoch]
        }

        # Evaluate on test set
        test_metrics = model.evaluate(test_dataset)
        best_metrics['test_loss'] = test_metrics[0]
        best_metrics['test_iou'] = test_metrics[1]
        best_metrics['test_dice'] = test_metrics[2]
        best_metrics['test_dice_nobg'] = test_metrics[3]
        best_metrics['test_dice_2'] = test_metrics[5]

        performance_summary = ({
            "Test Fold": test_fold,
            "Validation Fold": val_fold,
            "Train Folds": train_folds,
            "Best Epoch": best_epoch + 1,
            'Train Loss': history.history['loss'][best_epoch],
            'Train IoU': history.history['IoU'][best_epoch],
            "Final Train DSC": history.history['dsc'][best_epoch],
            "Final Train DSC no BG": history.history['dsc_nobg'][best_epoch],
            'Final Train DSC 1': history.history['dsc1'][best_epoch],
            'Final Train DSC 2': history.history['dsc2'][best_epoch],
            'Validation Loss': history.history['val_loss'][best_epoch],
            'Validation IoU': history.history['val_IoU'][best_epoch],
            "Final Validation DSC": history.history['val_dsc'][best_epoch],
            "Final Validation DSC no BG": history.history['val_dsc_nobg'][best_epoch],
            'Final Validation DSC 1': history.history['val_dsc1'][best_epoch],
            'Final Validation DSC 2': history.history['val_dsc2'][best_epoch],
            'Test Loss': test_metrics[0],
            'Test IoU': test_metrics[1],
            'Test Dice': test_metrics[2],
            'Test Dice no BG': test_metrics[3],
            'Test Dice 1': test_metrics[4],
            'Test Dice 2': test_metrics[5],
        })

        # Clear the model from memory
        del model
        
        # Clear datasets
        del train_dataset, val_dataset, test_dataset
        
    # Clear TensorFlow's session/graph
    tf.keras.backend.clear_session()
    
    # Force garbage collection
    import gc
    gc.collect()

    # Guardar en un archivo Excel en la fila excel_row
    output_file = "training_metrics.xlsx"

    # Actualizar el archivo Excel con el resumen de rendimiento
    update_excel(output_file, excel_row, performance_summary)